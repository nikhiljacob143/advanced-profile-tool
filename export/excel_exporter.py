# -*- coding: utf-8 -*-
# Copyright (C) 2026 Nikhil Jacob — GPL v2 or later
"""Excel workbook exporter built on openpyxl.

One workbook consolidates the run outputs: a Summary sheet (metadata,
volume totals and generation parameters) plus Long Section, Cross
Sections, Comparison and Volumes sheets as available.

Two modes:

* consolidated (default) — write-only (streaming) workbook; memory stays
  bounded for large sample counts. Freeze panes are not available in
  write-only workbooks, so header rows are bolded but not frozen.
* per-section — settings key ``excel_per_section`` true AND at most 200
  cross-sections: a normal workbook with one sheet per section
  (sanitised unique sheet names, frozen header row, 0.00/0.000 number
  formats) and, when ``excel_thumbnails`` is also true and a matching
  PNG exists in ``images_dir``, an embedded plot thumbnail to the right
  of the table.

The Summary sheet notes which mode was used. Offsets follow the plugin
convention: negative = left, positive = right. No qgis imports.
"""
import math
import os
import re

try:
    from openpyxl import Workbook
    from openpyxl.cell import WriteOnlyCell
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
    _BOLD = Font(bold=True)
except ImportError:                                    # pragma: no cover
    # openpyxl is an optional dependency; every other export format works
    # without it. export_workbook() raises a clear error when called.
    OPENPYXL_AVAILABLE = False
    Workbook = WriteOnlyCell = Font = get_column_letter = None
    _BOLD = None

from ..constants import DEFAULTS
from .image_exporter import build_image_name

__all__ = ["export_workbook", "OPENPYXL_AVAILABLE"]

#: User-facing message for the missing optional dependency.
OPENPYXL_MISSING_MSG = (
    "Excel export requires the Python package 'openpyxl', which is not "
    "installed in this QGIS environment. Install it (e.g. via the OSGeo4W "
    "shell or pip: python -m pip install openpyxl) and restart QGIS, or "
    "use the CSV export instead."
)

# per-section mode is refused above this section count (sheet bloat)
MAX_PER_SECTION_SHEETS = 200

_INVALID_SHEET_CHARS = re.compile(r"[\\/*?:\[\]]")


def _setting(settings, key, fallback=None):
    """Read ``key`` from a settings mapping/manager with factory fallback."""
    default = DEFAULTS.get(key, fallback)
    if settings is None:
        return default
    try:
        value = settings.get(key, default)
    except (TypeError, AttributeError):
        return default
    return default if value is None else value


def _decimals(settings):
    """Return (chainage, elevation, offset) decimal places."""
    return (int(_setting(settings, "decimals_chainage", 2)),
            int(_setting(settings, "decimals_elevation", 3)),
            int(_setting(settings, "decimals_offset", 2)))


def _num(value, decimals):
    """Round to fixed decimals; NaN/None → None (blank cell)."""
    if value is None:
        return None
    v = float(value)
    if math.isnan(v):
        return None
    return round(v, decimals)


def _header_row(ws, titles):
    """Append a bold header row to a write-only worksheet."""
    cells = []
    for title in titles:
        c = WriteOnlyCell(ws, value=title)
        c.font = _BOLD
        cells.append(c)
    ws.append(cells)


def _enabled(dem_defs):
    """Return enabled DEM definitions in given order."""
    return [d for d in dem_defs if getattr(d, "enabled", True)]


def _dem_cells(result, index, dems, dec_elev):
    """Per-DEM elevation cells for one sample of a ProfileResult."""
    out = []
    for dem in dems:
        line = result.lines.get(dem.layer_id)
        if line is None or line.elevations is None:
            out.append(None)
        else:
            out.append(_num(line.elevations[index], dec_elev))
    return out


def _summary_sheet(ws, run, settings, dec_elev):
    """Populate the Summary sheet: metadata, totals and parameters."""
    _header_row(ws, ["Advanced Profile Tool — export summary"])
    metadata = run.get("metadata") or {}
    alignment = run.get("alignment")
    if alignment is not None:
        metadata = dict(metadata)
        metadata.setdefault("Alignment", getattr(alignment, "name", ""))
        metadata.setdefault("Alignment CRS",
                            getattr(alignment, "crs_authid", ""))
        metadata.setdefault("Alignment length (m)",
                            round(getattr(alignment, "length", 0.0), 2))
    for key, value in metadata.items():
        ws.append([str(key), value])
    ws.append([])

    volumes = run.get("volumes")
    if volumes:
        _rows, totals = volumes
        _header_row(ws, ["Volume totals"])
        ws.append(["Total cut (m3)", _num((totals or {}).get("cut"), dec_elev)])
        ws.append(["Total fill (m3)", _num((totals or {}).get("fill"), dec_elev)])
        ws.append(["Net (fill - cut) (m3)",
                   _num((totals or {}).get("net"), dec_elev)])
        ws.append([])

    dems = _enabled(run.get("dem_defs") or [])
    if dems:
        _header_row(ws, ["Surfaces", "Band", "Interpolation",
                         "Vertical offset", "Reference"])
        for d in dems:
            ws.append([d.name, d.band, d.interp, d.v_offset,
                       "yes" if d.is_reference else "no"])
        ws.append([])

    _header_row(ws, ["Generation parameters"])
    if settings is not None:
        keys = ("chainage_format", "section_interval", "left_width",
                "right_width", "sampling_interval", "interp_method",
                "nodata_mode", "vertical_exaggeration",
                "decimals_chainage", "decimals_elevation", "decimals_offset")
        for key in keys:
            ws.append([key, _setting(settings, key)])


def _long_section_sheet(ws, result, dems, decs):
    """Populate the Long Section sheet (chainage, x, y, per-DEM RLs)."""
    dec_ch, dec_elev, dec_off = decs
    _header_row(ws, ["chainage", "x", "y"] + [d.name for d in dems])
    n = len(result.offsets) if result.offsets is not None else 0
    for i in range(n):
        ws.append([_num(result.offsets[i], dec_ch),
                   _num(result.xs[i], dec_off),
                   _num(result.ys[i], dec_off)]
                  + _dem_cells(result, i, dems, dec_elev))


def _cross_sections_sheet(ws, results, dems, decs):
    """Populate the consolidated Cross Sections sheet."""
    dec_ch, dec_elev, dec_off = decs
    _header_row(ws, ["section_label", "chainage", "offset", "x", "y"]
                + [d.name for d in dems])
    for result in results:
        n = len(result.offsets) if result.offsets is not None else 0
        ch = _num(result.chainage, dec_ch)
        for i in range(n):
            ws.append([result.label, ch,
                       _num(result.offsets[i], dec_off),
                       _num(result.xs[i], dec_off),
                       _num(result.ys[i], dec_off)]
                      + _dem_cells(result, i, dems, dec_elev))


def _comparison_sheet(ws, comparisons, decs):
    """Populate the Comparison sheet (per-section cut/fill areas)."""
    dec_ch, dec_elev, dec_off = decs
    _header_row(ws, ["label", "chainage", "cut_area", "fill_area",
                     "net_area", "gap_length", "valid"])
    for c in comparisons:
        ws.append([c.label, _num(c.chainage, dec_ch),
                   _num(c.cut_area, dec_elev), _num(c.fill_area, dec_elev),
                   _num(c.net_area, dec_elev), _num(c.gap_length, dec_off),
                   "yes" if c.valid else "no"])


def _volumes_sheet(ws, rows, totals, decs):
    """Populate the Volumes sheet with a totals row at the end."""
    dec_ch, dec_elev, _dec_off = decs
    _header_row(ws, ["from_label", "to_label", "from_chainage",
                     "to_chainage", "length", "cut_volume", "fill_volume",
                     "net_volume", "cum_cut", "cum_fill", "cum_net"])
    for r in rows:
        ws.append([getattr(r, "from_label", "") or f"S{r.from_id}",
                   getattr(r, "to_label", "") or f"S{r.to_id}",
                   _num(r.from_chainage, dec_ch),
                   _num(r.to_chainage, dec_ch),
                   _num(r.length, dec_ch),
                   _num(r.cut_volume, dec_elev),
                   _num(r.fill_volume, dec_elev),
                   _num(r.net_volume, dec_elev),
                   _num(r.cum_cut, dec_elev),
                   _num(r.cum_fill, dec_elev),
                   _num(r.cum_net, dec_elev)])
    totals = totals or {}
    total_cells = ["TOTAL", None, None, None, None,
                   _num(totals.get("cut", 0.0), dec_elev),
                   _num(totals.get("fill", 0.0), dec_elev),
                   _num(totals.get("net", 0.0), dec_elev),
                   None, None, None]
    styled = []
    for v in total_cells:
        c = WriteOnlyCell(ws, value=v)
        c.font = _BOLD
        styled.append(c)
    ws.append(styled)


def _sheet_name(label, used):
    """Sanitise ``label`` into a valid, unique Excel sheet name (max 31
    characters; \\ / * ? : [ ] removed)."""
    base = _INVALID_SHEET_CHARS.sub("_", str(label)).strip() or "Section"
    base = base[:31]
    name = base
    n = 1
    while name.lower() in used:
        suffix = f"_{n}"
        name = base[:31 - len(suffix)] + suffix
        n += 1
    used.add(name.lower())
    return name


def _section_sheet(ws, result, dems, decs):
    """Populate one per-section sheet: offset, x, y, per-DEM RLs with a
    frozen header row and fixed number formats (0.00 / 0.000)."""
    dec_ch, dec_elev, dec_off = decs
    _header_row(ws, ["offset (m)", "x (m)", "y (m)"]
                + [f"{d.name} RL (m)" for d in dems])
    n = len(result.offsets) if result.offsets is not None else 0
    for i in range(n):
        ws.append([_num(result.offsets[i], dec_off),
                   _num(result.xs[i], dec_off),
                   _num(result.ys[i], dec_off)]
                  + _dem_cells(result, i, dems, dec_elev))
    ws.freeze_panes = "A2"
    max_col = 3 + len(dems)
    for row in ws.iter_rows(min_row=2, min_col=1, max_col=max_col):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.number_format = "0.000" if cell.column > 3 else "0.00"


def _embed_thumbnail(ws, result, dems, images_dir):
    """Embed the section's plot PNG (named per
    image_exporter.build_image_name) to the right of the table. Silently
    skipped when the image or Pillow is unavailable."""
    png = os.path.join(str(images_dir), build_image_name(result.label))
    if not os.path.exists(png):
        return False
    try:
        from openpyxl.drawing.image import Image as XlImage
        anchor = f"{get_column_letter(3 + len(dems) + 2)}2"
        ws.add_image(XlImage(png), anchor)
        return True
    except Exception:                     # Pillow missing / unreadable PNG
        return False


def _export_per_section(path, run, settings, decs, dems, images_dir):
    """Per-section workbook (normal mode): Summary + shared sheets + one
    sheet per cross-section, optionally with embedded thumbnails."""
    thumbnails = bool(_setting(settings, "excel_thumbnails", False))
    cross_sections = run.get("cross_sections") or []
    wb = Workbook()
    try:
        wb.remove(wb.active)              # drop the default "Sheet"
        summary = wb.create_sheet("Summary")
        _summary_sheet(summary, run, settings, decs[1])
        summary.append([])
        summary.append(["Export mode",
                        "per-section sheets"
                        + (" with thumbnails" if thumbnails else "")])

        long_section = run.get("long_section")
        if long_section is not None:
            _long_section_sheet(wb.create_sheet("Long Section"),
                                long_section, dems, decs)

        comparisons = run.get("comparisons")
        if comparisons:
            _comparison_sheet(wb.create_sheet("Comparison"),
                              comparisons, decs)

        volumes = run.get("volumes")
        if volumes:
            rows, totals = volumes
            _volumes_sheet(wb.create_sheet("Volumes"), rows, totals, decs)

        used = {"summary", "long section", "comparison", "volumes"}
        for result in cross_sections:
            ws = wb.create_sheet(_sheet_name(result.label, used))
            _section_sheet(ws, result, dems, decs)
            if thumbnails and images_dir:
                _embed_thumbnail(ws, result, dems, images_dir)

        wb.save(path)
    finally:
        wb.close()
    return path


def _export_streaming(path, run, settings, decs, dems):
    """Consolidated write-only workbook (original streaming path)."""
    wb = Workbook(write_only=True)
    try:
        summary = wb.create_sheet("Summary")
        _summary_sheet(summary, run, settings, decs[1])
        summary.append([])
        summary.append(["Export mode", "consolidated (streaming)"])

        long_section = run.get("long_section")
        if long_section is not None:
            _long_section_sheet(wb.create_sheet("Long Section"),
                                long_section, dems, decs)

        cross_sections = run.get("cross_sections")
        if cross_sections:
            _cross_sections_sheet(wb.create_sheet("Cross Sections"),
                                  cross_sections, dems, decs)

        comparisons = run.get("comparisons")
        if comparisons:
            _comparison_sheet(wb.create_sheet("Comparison"),
                              comparisons, decs)

        volumes = run.get("volumes")
        if volumes:
            rows, totals = volumes
            _volumes_sheet(wb.create_sheet("Volumes"), rows, totals, decs)

        wb.save(path)
    finally:
        wb.close()
    return path


def export_workbook(path, run, settings, images_dir=None):
    """Write the run's Excel workbook and return the path.

    ``run`` is a dict that may contain: ``long_section`` (ProfileResult),
    ``cross_sections`` (list of ProfileResult), ``comparisons`` (list of
    SectionComparison), ``volumes`` ((rows, totals) tuple), ``dem_defs``
    (list of DemDef), ``alignment`` (AlignmentDef) and ``metadata`` (dict
    shown on the Summary sheet). Missing keys simply omit their sheet.

    Mode selection: when settings key ``excel_per_section`` is true AND
    there are 1–200 cross-sections, a normal workbook is written with
    one sheet per section (frozen header, 0.00/0.000 number formats) and
    optional embedded thumbnails (settings key ``excel_thumbnails``;
    PNGs looked up in ``images_dir`` by
    ``image_exporter.build_image_name(section label)``). Otherwise the
    original write-only consolidated workbook is produced. The Summary
    sheet records the mode used. Numbers are rounded per the plugin
    decimal settings.

    Raises ``RuntimeError`` with an installation hint when the optional
    ``openpyxl`` dependency is not available.
    """
    if not OPENPYXL_AVAILABLE:
        raise RuntimeError(OPENPYXL_MISSING_MSG)
    decs = _decimals(settings)
    dems = _enabled(run.get("dem_defs") or [])
    cross_sections = run.get("cross_sections") or []
    per_section = bool(_setting(settings, "excel_per_section", False))
    if per_section and 0 < len(cross_sections) <= MAX_PER_SECTION_SHEETS:
        return _export_per_section(path, run, settings, decs, dems,
                                   images_dir)
    return _export_streaming(path, run, settings, decs, dems)
